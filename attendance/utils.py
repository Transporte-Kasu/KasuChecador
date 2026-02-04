from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime, timedelta, date
from .models import (
    Asistencia, TipoMovimiento, Empleado, ConfiguracionSistema, TiempoExtra, TipoHorario,
    HorarioDiaSemana, AsignacionTurnoRotativo, TipoSistemaHorario
)
import os
from django.conf import settings

def obtener_horario_esperado(empleado, fecha):
    """
    Obtiene el horario esperado de un empleado para una fecha específica.
    
    Args:
        empleado: Instancia de Empleado
        fecha: datetime.date o datetime.datetime
    
    Returns:
        dict: {
            'hora_entrada': time o None,
            'hora_salida': time o None,
            'es_dia_laboral': bool,
            'tolerancia_minutos': int,
            'tiene_horario_comida': bool,
            'hora_inicio_comida': time o None,
            'hora_fin_comida': time o None,
            'tipo_sistema': str
        }
    """
    # Convertir datetime a date si es necesario
    if isinstance(fecha, datetime):
        fecha = fecha.date()
    
    tipo_horario = empleado.tipo_horario
    
    # Si no tiene tipo de horario asignado, usar configuración global
    if not tipo_horario:
        config = ConfiguracionSistema.objects.first()
        if config:
            return {
                'hora_entrada': config.hora_entrada,
                'hora_salida': None,
                'es_dia_laboral': fecha.weekday() < 5,  # Lunes a viernes
                'tolerancia_minutos': config.minutos_tolerancia,
                'tiene_horario_comida': False,
                'hora_inicio_comida': None,
                'hora_fin_comida': None,
                'tipo_sistema': 'FIJO'
            }
        else:
            # Fallback si no hay configuración
            return {
                'hora_entrada': datetime.strptime('09:00:00', '%H:%M:%S').time(),
                'hora_salida': None,
                'es_dia_laboral': fecha.weekday() < 5,
                'tolerancia_minutos': 15,
                'tiene_horario_comida': False,
                'hora_inicio_comida': None,
                'hora_fin_comida': None,
                'tipo_sistema': 'FIJO'
            }
    
    # Según el tipo de sistema
    tipo_sistema = tipo_horario.tipo_sistema
    
    # TURNO 24x24 HORAS
    if tipo_sistema == TipoSistemaHorario.TURNO_24H or tipo_horario.es_turno_24h:
        # Para turnos de 24h, el día laboral depende del ciclo
        # Este tipo de horario se maneja de forma especial en calcular_retardo
        return {
            'hora_entrada': tipo_horario.hora_entrada,
            'hora_salida': tipo_horario.hora_salida,
            'es_dia_laboral': True,  # Siempre es laboral, pero cada 48h
            'tolerancia_minutos': tipo_horario.minutos_tolerancia,
            'tiene_horario_comida': False,  # Turnos de 24h no tienen comida formal
            'hora_inicio_comida': None,
            'hora_fin_comida': None,
            'tipo_sistema': 'TURNO_24H'
        }
    
    # HORARIO ROTATIVO
    elif tipo_sistema == TipoSistemaHorario.ROTATIVO:
        # Buscar la asignación de turno activa para la fecha
        asignacion = AsignacionTurnoRotativo.objects.filter(
            empleado=empleado,
            fecha_inicio__lte=fecha,
            fecha_fin__gte=fecha,
            activo=True
        ).select_related('turno_rotativo').first()
        
        if asignacion:
            turno = asignacion.turno_rotativo
            return {
                'hora_entrada': turno.hora_entrada,
                'hora_salida': turno.hora_salida,
                'es_dia_laboral': True,
                'tolerancia_minutos': tipo_horario.minutos_tolerancia,
                'tiene_horario_comida': tipo_horario.tiene_horario_comida,
                'hora_inicio_comida': tipo_horario.hora_inicio_comida,
                'hora_fin_comida': tipo_horario.hora_fin_comida,
                'tipo_sistema': 'ROTATIVO'
            }
        else:
            # No hay asignación para esta fecha, usar horario base
            return {
                'hora_entrada': tipo_horario.hora_entrada,
                'hora_salida': tipo_horario.hora_salida,
                'es_dia_laboral': False,  # No tiene turno asignado
                'tolerancia_minutos': tipo_horario.minutos_tolerancia,
                'tiene_horario_comida': tipo_horario.tiene_horario_comida,
                'hora_inicio_comida': tipo_horario.hora_inicio_comida,
                'hora_fin_comida': tipo_horario.hora_fin_comida,
                'tipo_sistema': 'ROTATIVO'
            }
    
    # HORARIO PERSONALIZADO POR DÍA
    elif tipo_sistema == TipoSistemaHorario.PERSONALIZADO or tipo_horario.requiere_horario_por_dia:
        # Buscar configuración para el día de la semana
        dia_semana = fecha.weekday()  # 0=Lunes, 6=Domingo
        
        horario_dia = HorarioDiaSemana.objects.filter(
            tipo_horario=tipo_horario,
            dia_semana=dia_semana
        ).first()
        
        if horario_dia:
            return {
                'hora_entrada': horario_dia.hora_entrada,
                'hora_salida': horario_dia.hora_salida,
                'es_dia_laboral': horario_dia.es_dia_laboral,
                'tolerancia_minutos': tipo_horario.minutos_tolerancia,
                'tiene_horario_comida': bool(horario_dia.hora_inicio_comida and horario_dia.hora_fin_comida),
                'hora_inicio_comida': horario_dia.hora_inicio_comida,
                'hora_fin_comida': horario_dia.hora_fin_comida,
                'tipo_sistema': 'PERSONALIZADO'
            }
        else:
            # No hay configuración para este día, usar horario base o marcar como no laboral
            return {
                'hora_entrada': tipo_horario.hora_entrada,
                'hora_salida': tipo_horario.hora_salida,
                'es_dia_laboral': fecha.weekday() < 5,  # Default: lunes a viernes
                'tolerancia_minutos': tipo_horario.minutos_tolerancia,
                'tiene_horario_comida': tipo_horario.tiene_horario_comida,
                'hora_inicio_comida': tipo_horario.hora_inicio_comida,
                'hora_fin_comida': tipo_horario.hora_fin_comida,
                'tipo_sistema': 'PERSONALIZADO'
            }
    
    # HORARIO FIJO (default)
    else:
        return {
            'hora_entrada': tipo_horario.hora_entrada,
            'hora_salida': tipo_horario.hora_salida,
            'es_dia_laboral': fecha.weekday() < 5,  # Lunes a viernes por defecto
            'tolerancia_minutos': tipo_horario.minutos_tolerancia,
            'tiene_horario_comida': tipo_horario.tiene_horario_comida,
            'hora_inicio_comida': tipo_horario.hora_inicio_comida,
            'hora_fin_comida': tipo_horario.hora_fin_comida,
            'tipo_sistema': 'FIJO'
        }

def enviar_email_visitante(visitante):
    """Envía email con QR al visitante y notifica al departamento"""

    # Email al visitante
    subject_visitante = f'Confirmación de Visita - {visitante.fecha_visita}'

    html_message = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background-color: #3b82f6; padding: 20px; text-align: center;">
            <h1 style="color: white; margin: 0;">Visita Confirmada</h1>
        </div>
        <div style="padding: 20px; background-color: #f9fafb;">
            <p>Hola <strong>{visitante.nombre}</strong>,</p>
            <p>Tu visita ha sido confirmada con los siguientes detalles:</p>
            <div style="background-color: white; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <p><strong>Departamento:</strong> {visitante.departamento_visita.nombre}</p>
                <p><strong>Fecha:</strong> {visitante.fecha_visita.strftime('%d/%m/%Y')}</p>
                <p><strong>Hora:</strong> {visitante.hora_visita.strftime('%H:%M')}</p>
                <p><strong>Motivo:</strong> {visitante.motivo}</p>
            </div>
            <div style="text-align: center; margin: 30px 0;">
                <p><strong>Tu código QR de acceso:</strong></p>
                <img src="{visitante.qr_code.url}" alt="QR Code" style="max-width: 250px;">
                <p style="font-size: 12px; color: #6b7280;">Presenta este código al llegar a recepción</p>
            </div>
        </div>
    </body>
    </html>
    """

    # Usar EmailMultiAlternatives en lugar de EmailMessage
    email_visitante = EmailMultiAlternatives(
        subject_visitante,
        'Tu visita ha sido confirmada. Por favor revisa el contenido HTML del email.',
        settings.DEFAULT_FROM_EMAIL,
        [visitante.email]
    )
    email_visitante.attach_alternative(html_message, "text/html")

    # Adjuntar QR
    #if visitante.qr_code:
    #    with open(visitante.qr_code.path, 'rb') as f:
    #        email_visitante.attach('qr_code.png', f.read(), 'image/png')

    email_visitante.send(fail_silently=False)

    # Email al departamento
    subject_depto = f'Nueva Visita Programada - {visitante.nombre}'
    mensaje_depto = f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <h2>Nueva Visita Programada</h2>
        <p>Se ha programado una visita para su departamento:</p>
        <ul>
            <li><strong>Visitante:</strong> {visitante.nombre}</li>
            <li><strong>Empresa:</strong> {visitante.empresa or 'N/A'}</li>
            <li><strong>Fecha:</strong> {visitante.fecha_visita.strftime('%d/%m/%Y')}</li>
            <li><strong>Hora:</strong> {visitante.hora_visita.strftime('%H:%M')}</li>
            <li><strong>Motivo:</strong> {visitante.motivo}</li>
        </ul>
    </body>
    </html>
    """

    email_depto = EmailMultiAlternatives(
        subject_depto,
        'Nueva visita programada. Por favor revisa el contenido HTML del email.',
        settings.DEFAULT_FROM_EMAIL,
        [visitante.departamento_visita.email]
    )
    email_depto.attach_alternative(mensaje_depto, "text/html")
    email_depto.send(fail_silently=False)


def generar_reporte_semanal():
    """Genera y envía el reporte semanal todos los jueves"""
    hoy = timezone.now().date()

    # Calcular el rango de la semana (lunes a jueves)
    # Si hoy es jueves (weekday = 3), la semana va desde el lunes anterior hasta hoy
    dias_desde_lunes = hoy.weekday()  # 0=lunes, 3=jueves
    fecha_inicio = hoy - timedelta(days=dias_desde_lunes)
    fecha_fin = hoy

    # Obtener configuración
    config = ConfiguracionSistema.objects.first()
    if not config:
        return

    # Obtener datos por empleado
    empleados = Empleado.objects.filter(activo=True)

    html_reporte = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: center; }}
            th {{ background-color: #3b82f6; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .titulo {{ background-color: #1e40af; color: white; padding: 20px; text-align: center; }}
            .resumen {{ background-color: #eff6ff; padding: 20px; border-radius: 8px; margin: 20px 0; }}
            .alerta {{ background-color: #fef2f2; padding: 15px; border-left: 4px solid #ef4444; margin: 20px 0; }}
        </style>
    </head>
    <body>
        <div class="titulo">
            <h1>Reporte Semanal de Asistencias</h1>
            <p>{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}</p>
        </div>

        <table>
            <tr>
                <th>Empleado</th>
                <th>Código</th>
                <th>Departamento</th>
                <th>Días Asistidos</th>
                <th>Retardos</th>
                <th>Total Min. Retardo</th>
                <th>Faltas</th>
            </tr>
    """

    # Recolectar empleados con retardos consecutivos
    empleados_retardos_consecutivos = []

    for empleado in empleados:
        asistencias = Asistencia.objects.filter(
            empleado=empleado,
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin,
            tipo_movimiento=TipoMovimiento.ENTRADA
        )

        dias_asistidos = asistencias.values('fecha').distinct().count()
        retardos = asistencias.filter(retardo=True).count()
        total_min_retardo = sum(asistencias.filter(retardo=True).values_list('minutos_retardo', flat=True))

        # Calcular días/turnos laborales según tipo de horario
        tipo_horario = empleado.tipo_horario
        if tipo_horario and tipo_horario.es_turno_24h:
            # Para turnos de 24h: calcular turnos esperados en el período
            # Ciclo de 48 horas (24h trabajo + 24h descanso)
            dias_periodo = (fecha_fin - fecha_inicio).days + 1
            turnos_esperados = dias_periodo // 2  # Un turno cada 2 días
            faltas = max(0, turnos_esperados - dias_asistidos)
        else:
            # Para horarios regulares: lunes a viernes
            dias_laborales = 0
            fecha_actual = fecha_inicio
            while fecha_actual <= fecha_fin:
                if fecha_actual.weekday() < 5:  # Lunes a viernes
                    dias_laborales += 1
                fecha_actual += timedelta(days=1)
            faltas = dias_laborales - dias_asistidos

        html_reporte += f"""
            <tr>
                <td>{empleado.user.get_full_name()}</td>
                <td>{empleado.codigo_empleado}</td>
                <td>{empleado.departamento.nombre if empleado.departamento else 'N/A'}</td>
                <td>{dias_asistidos}</td>
                <td>{retardos}</td>
                <td>{total_min_retardo}</td>
                <td>{faltas}</td>
            </tr>
        """

        # Detectar empleados con retardos consecutivos (3 o más retardos en la semana)
        if retardos >= 3:
            empleados_retardos_consecutivos.append({
                'nombre': empleado.user.get_full_name(),
                'codigo': empleado.codigo_empleado,
                'retardos': retardos
            })

    html_reporte += "</table>"

    # Agregar alerta de retardos consecutivos si existen
    if empleados_retardos_consecutivos:
        html_reporte += """
        <div class="alerta">
            <h2>⚠️ Atención: Retardos Recurrentes</h2>
            <p>Los siguientes empleados tienen 3 o más retardos esta semana:</p>
            <table>
                <tr>
                    <th>Empleado</th>
                    <th>Código</th>
                    <th>Retardos (esta semana)</th>
                </tr>
        """

        for emp in empleados_retardos_consecutivos:
            html_reporte += f"""
                <tr>
                    <td>{emp['nombre']}</td>
                    <td>{emp['codigo']}</td>
                    <td>{emp['retardos']}</td>
                </tr>
            """

        html_reporte += "</table></div>"

    html_reporte += "</body></html>"

    # Generar archivo Excel
    excel_buffer = generar_excel_reporte_semanal(fecha_inicio, fecha_fin)
    nombre_excel = f"reporte_semanal_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"

    # Enviar email
    email = EmailMultiAlternatives(
        f'Reporte Semanal de Asistencias - Semana del {fecha_inicio.strftime("%d/%m/%Y")}',
        'Reporte semanal de asistencias. Por favor revisa el contenido HTML y el archivo Excel adjunto con el detalle de todas las checadas.',
        settings.DEFAULT_FROM_EMAIL,
        [config.email_gerente,'zuly.becerra@loginco.com.mx']
    )
    email.attach_alternative(html_reporte, "text/html")
    
    # Adjuntar archivo Excel
    email.attach(nombre_excel, excel_buffer.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    email.send(fail_silently=False)


def generar_reporte_diario():
    """Genera y envía el reporte diario después de las 12:00 PM"""
    hoy = timezone.now().date()

    # Obtener configuración
    config = ConfiguracionSistema.objects.first()
    if not config:
        return

    # Asistencias del día
    asistencias_entrada = Asistencia.objects.filter(
        fecha=hoy,
        tipo_movimiento=TipoMovimiento.ENTRADA
    ).select_related('empleado', 'empleado__user')

    total_empleados = Empleado.objects.filter(activo=True).count()
    llegaron = asistencias_entrada.count()
    retardos = asistencias_entrada.filter(retardo=True)

    # Empleados con retardos consecutivos (últimos 5 días)
    fecha_inicio = hoy - timedelta(days=5)
    empleados_retardos_consecutivos = []

    for empleado in Empleado.objects.filter(activo=True):
        retardos_count = Asistencia.objects.filter(
            empleado=empleado,
            fecha__gte=fecha_inicio,
            fecha__lte=hoy,
            tipo_movimiento=TipoMovimiento.ENTRADA,
            retardo=True
        ).count()

        if retardos_count >= 3:
            empleados_retardos_consecutivos.append({
                'nombre': empleado.user.get_full_name(),
                'codigo': empleado.codigo_empleado,
                'retardos': retardos_count
            })

    # Generar HTML del reporte
    html_reporte = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
            th {{ background-color: #3b82f6; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .resumen {{ background-color: #eff6ff; padding: 20px; border-radius: 8px; margin: 20px 0; }}
            .alerta {{ background-color: #fef2f2; padding: 15px; border-left: 4px solid #ef4444; margin: 20px 0; }}
        </style>
    </head>
    <body>
        <h1>Reporte Diario de Asistencia</h1>
        <p><strong>Fecha:</strong> {hoy.strftime('%d/%m/%Y')}</p>

        <div class="resumen">
            <h2>Resumen</h2>
            <p><strong>Total de Empleados:</strong> {total_empleados}</p>
            <p><strong>Asistieron:</strong> {llegaron} ({(llegaron/total_empleados*100):.1f}%)</p>
            <p><strong>Retardos del Día:</strong> {retardos.count()}</p>
        </div>

        <h2>Retardos del Día</h2>
        <table>
            <tr>
                <th>Empleado</th>
                <th>Código</th>
                <th>Tipo de Horario</th>
                <th>Hora de Entrada</th>
                <th>Minutos de Retardo</th>
            </tr>
    """

    for asistencia in retardos:
        tipo_horario_nombre = asistencia.empleado.tipo_horario.nombre if asistencia.empleado.tipo_horario else 'Estándar'
        html_reporte += f"""
            <tr>
                <td>{asistencia.empleado.user.get_full_name()}</td>
                <td>{asistencia.empleado.codigo_empleado}</td>
                <td>{tipo_horario_nombre}</td>
                <td>{asistencia.hora.strftime('%H:%M')}</td>
                <td>{asistencia.minutos_retardo}</td>
            </tr>
        """

    html_reporte += "</table>"

    if empleados_retardos_consecutivos:
        html_reporte += """
        <div class="alerta">
            <h2>⚠️ Atención: Retardos Consecutivos</h2>
            <p>Los siguientes empleados tienen 3 o más retardos en los últimos 5 días:</p>
            <table>
                <tr>
                    <th>Empleado</th>
                    <th>Código</th>
                    <th>Retardos (últimos 5 días)</th>
                </tr>
        """

        for emp in empleados_retardos_consecutivos:
            html_reporte += f"""
                <tr>
                    <td>{emp['nombre']}</td>
                    <td>{emp['codigo']}</td>
                    <td>{emp['retardos']}</td>
                </tr>
            """

        html_reporte += "</table></div>"

    html_reporte += "</body></html>"

    # Enviar email
    email = EmailMultiAlternatives(
        f'Reporte Diario de Asistencia - {hoy.strftime("%d/%m/%Y")}',
        'Reporte diario de asistencias. Por favor revisa el contenido HTML.',
        settings.DEFAULT_FROM_EMAIL,
        [config.email_gerente,'zuly.becerra@loginco.com.mx']
    )
    email.attach_alternative(html_reporte, "text/html")
    email.send(fail_silently=False)


def generar_reporte_quincenal(dia):
    """Genera el reporte quincenal (días 13 y 28)"""
    hoy = timezone.now().date()

    # Determinar el período
    if dia == 13:
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy.replace(day=13)
        periodo = "Primera Quincena"
    else:  # día 28
        fecha_inicio = hoy.replace(day=14)
        # Último día del mes
        if hoy.month == 12:
            fecha_fin = hoy.replace(day=31)
        else:
            fecha_fin = (hoy.replace(month=hoy.month + 1, day=1) - timedelta(days=1))
        periodo = "Segunda Quincena"

    config = ConfiguracionSistema.objects.first()
    if not config:
        return

    # Obtener datos por empleado
    empleados = Empleado.objects.filter(activo=True)

    html_reporte = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: center; }}
            th {{ background-color: #3b82f6; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .titulo {{ background-color: #1e40af; color: white; padding: 20px; text-align: center; }}
        </style>
    </head>
    <body>
        <div class="titulo">
            <h1>Reporte de Asistencias - {periodo}</h1>
            <p>{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}</p>
        </div>

        <table>
            <tr>
                <th>Empleado</th>
                <th>Código</th>
                <th>Departamento</th>
                <th>Días Asistidos</th>
                <th>Retardos</th>
                <th>Total Min. Retardo</th>
                <th>Faltas</th>
            </tr>
    """

    for empleado in empleados:
        asistencias = Asistencia.objects.filter(
            empleado=empleado,
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin,
            tipo_movimiento=TipoMovimiento.ENTRADA
        )

        dias_asistidos = asistencias.values('fecha').distinct().count()
        retardos = asistencias.filter(retardo=True).count()
        total_min_retardo = sum(asistencias.filter(retardo=True).values_list('minutos_retardo', flat=True))

        # Calcular días/turnos laborales según tipo de horario
        tipo_horario = empleado.tipo_horario
        if tipo_horario and tipo_horario.es_turno_24h:
            # Para turnos de 24h: calcular turnos esperados en el período
            dias_periodo = (fecha_fin - fecha_inicio).days + 1
            turnos_esperados = dias_periodo // 2
            faltas = max(0, turnos_esperados - dias_asistidos)
        else:
            # Para horarios regulares: todos los días del período
            dias_laborales = (fecha_fin - fecha_inicio).days + 1
            faltas = dias_laborales - dias_asistidos

        html_reporte += f"""
            <tr>
                <td>{empleado.user.get_full_name()}</td>
                <td>{empleado.codigo_empleado}</td>
                <td>{empleado.departamento.nombre if empleado.departamento else 'N/A'}</td>
                <td>{dias_asistidos}</td>
                <td>{retardos}</td>
                <td>{total_min_retardo}</td>
                <td>{faltas}</td>
            </tr>
        """

    html_reporte += "</table></body></html>"

    # Enviar email
    email = EmailMultiAlternatives(
        f'Reporte Quincenal - {periodo} - {hoy.strftime("%B %Y")}',
        'Reporte quincenal de asistencias. Por favor revisa el contenido HTML.',
        settings.DEFAULT_FROM_EMAIL,
        [config.email_gerente]
    )
    email.attach_alternative(html_reporte, "text/html")
    email.send(fail_silently=False)


def generar_reporte_tiempo_extra_mensual():
    """Genera el reporte mensual de tiempo extra y lo guarda en la red"""
    hoy = timezone.now()
    mes = hoy.month
    anio = hoy.year

    config = ConfiguracionSistema.objects.first()
    if not config or not config.ruta_red_reportes:
        return

    # Obtener tiempos extra del mes
    tiempos_extra = TiempoExtra.objects.filter(
        fecha__month=mes,
        fecha__year=anio,
        aprobado=True
    ).select_related('empleado', 'empleado__user')

    # Generar HTML
    html_reporte = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
            th {{ background-color: #10b981; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .total {{ background-color: #d1fae5; font-weight: bold; }}
        </style>
    </head>
    <body>
        <h1>Reporte de Tiempo Extra</h1>
        <p><strong>Período:</strong> {hoy.strftime('%B %Y')}</p>

        <table>
            <tr>
                <th>Empleado</th>
                <th>Código</th>
                <th>Fecha</th>
                <th>Horas Extra</th>
                <th>Descripción</th>
            </tr>
    """

    total_horas = 0
    empleados_resumen = {}

    for te in tiempos_extra:
        html_reporte += f"""
            <tr>
                <td>{te.empleado.user.get_full_name()}</td>
                <td>{te.empleado.codigo_empleado}</td>
                <td>{te.fecha.strftime('%d/%m/%Y')}</td>
                <td>{te.horas_extra}</td>
                <td>{te.descripcion}</td>
            </tr>
        """
        total_horas += float(te.horas_extra)

        emp_id = te.empleado.id
        if emp_id not in empleados_resumen:
            empleados_resumen[emp_id] = {
                'nombre': te.empleado.user.get_full_name(),
                'codigo': te.empleado.codigo_empleado,
                'horas': 0
            }
        empleados_resumen[emp_id]['horas'] += float(te.horas_extra)

    html_reporte += f"""
            <tr class="total">
                <td colspan="3">TOTAL</td>
                <td>{total_horas:.2f}</td>
                <td></td>
            </tr>
        </table>

        <h2>Resumen por Empleado</h2>
        <table>
            <tr>
                <th>Empleado</th>
                <th>Código</th>
                <th>Total Horas Extra</th>
            </tr>
    """

    for emp in empleados_resumen.values():
        html_reporte += f"""
            <tr>
                <td>{emp['nombre']}</td>
                <td>{emp['codigo']}</td>
                <td>{emp['horas']:.2f}</td>
            </tr>
        """

    html_reporte += "</table></body></html>"

    # Guardar en ruta de red
    nombre_archivo = f"reporte_tiempo_extra_{anio}_{mes:02d}.html"
    ruta_completa = os.path.join(config.ruta_red_reportes, nombre_archivo)

    try:
        with open(ruta_completa, 'w', encoding='utf-8') as f:
            f.write(html_reporte)
        print(f"Reporte guardado en: {ruta_completa}")
    except Exception as e:
        print(f"Error al guardar reporte: {e}")

# ========== FUNCIONES PARA REPORTES EXCEL ==========

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def generar_excel_reporte_semanal(fecha_inicio, fecha_fin):
    """
    Genera un archivo Excel con todas las checadas de la semana.
    
    Args:
        fecha_inicio: datetime.date - Inicio del período (lunes)
        fecha_fin: datetime.date - Fin del período (jueves o viernes)
    
    Returns:
        BytesIO: Buffer con el archivo Excel generado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Semanal"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "REPORTE SEMANAL DE ASISTENCIAS"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    ws.merge_cells('A2:H2')
    
    # Encabezados
    headers = ['Fecha', 'Empleado', 'Código', 'Departamento', 'Entrada', 'Salida Comida', 'Entrada Comida', 'Salida']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Obtener empleados activos
    empleados = Empleado.objects.filter(activo=True).select_related('user', 'departamento').order_by('codigo_empleado')
    
    row = 5
    fecha_actual = fecha_inicio
    
    # Iterar por cada fecha en el rango
    while fecha_actual <= fecha_fin:
        for empleado in empleados:
            # Obtener asistencias del día
            asistencias = Asistencia.objects.filter(
                empleado=empleado,
                fecha=fecha_actual
            ).order_by('hora')
            
            if asistencias.exists():
                # Organizar por tipo de movimiento
                checadas = {
                    'ENTRADA': None,
                    'SALIDA_COMIDA': None,
                    'ENTRADA_COMIDA': None,
                    'SALIDA': None
                }
                
                for asist in asistencias:
                    checadas[asist.tipo_movimiento] = asist
                
                # Escribir fila
                ws.cell(row=row, column=1, value=fecha_actual.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=2, value=empleado.user.get_full_name())
                ws.cell(row=row, column=3, value=empleado.codigo_empleado)
                ws.cell(row=row, column=4, value=empleado.departamento.nombre if empleado.departamento else 'N/A')
                
                # Entrada
                if checadas['ENTRADA']:
                    hora_str = checadas['ENTRADA'].hora.strftime('%H:%M')
                    if checadas['ENTRADA'].retardo:
                        hora_str += f" (Retardo: {checadas['ENTRADA'].minutos_retardo} min)"
                    ws.cell(row=row, column=5, value=hora_str)
                
                # Salida a comida
                if checadas['SALIDA_COMIDA']:
                    ws.cell(row=row, column=6, value=checadas['SALIDA_COMIDA'].hora.strftime('%H:%M'))
                
                # Entrada de comida
                if checadas['ENTRADA_COMIDA']:
                    ws.cell(row=row, column=7, value=checadas['ENTRADA_COMIDA'].hora.strftime('%H:%M'))
                
                # Salida
                if checadas['SALIDA']:
                    ws.cell(row=row, column=8, value=checadas['SALIDA'].hora.strftime('%H:%M'))
                
                # Aplicar bordes
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = border
                
                row += 1
        
        fecha_actual += timedelta(days=1)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_mensual(mes, anio):
    """
    Genera un archivo Excel detallado del reporte mensual.
    
    Args:
        mes: int - Mes (1-12)
        anio: int - Año
    
    Returns:
        BytesIO: Buffer con el archivo Excel generado
    """
    wb = Workbook()
    
    # Hoja 1: Resumen por empleado
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Hoja 2: Detalle de asistencias
    ws_detalle = wb.create_sheet("Detalle de Asistencias")
    
    # Hoja 3: Retardos y faltas
    ws_retardos = wb.create_sheet("Retardos y Faltas")
    
    # Estilos
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Obtener nombre del mes
    from calendar import month_name
    import locale
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except:
        pass
    nombre_mes = month_name[mes] if mes <= 12 else str(mes)
    
    # ===== HOJA 1: RESUMEN =====
    ws_resumen['A1'] = f"REPORTE MENSUAL DE ASISTENCIAS - {nombre_mes.upper()} {anio}"
    ws_resumen['A1'].font = title_font
    ws_resumen.merge_cells('A1:H1')
    
    headers_resumen = ['Empleado', 'Código', 'Departamento', 'Días Asistidos', 'Retardos', 'Min. Retardo', 'Faltas', 'Permisos']
    for col, header in enumerate(headers_resumen, start=1):
        cell = ws_resumen.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Obtener datos
    from .models import SolicitudPermiso, SolicitudVacaciones, EstadoSolicitud
    empleados = Empleado.objects.filter(activo=True).select_related('user', 'departamento').order_by('codigo_empleado')
    
    # Calcular días laborales del mes
    from calendar import monthrange
    dias_mes = monthrange(anio, mes)[1]
    fecha_inicio_mes = date(anio, mes, 1)
    fecha_fin_mes = date(anio, mes, dias_mes)
    
    row_resumen = 4
    for empleado in empleados:
        asistencias = Asistencia.objects.filter(
            empleado=empleado,
            fecha__month=mes,
            fecha__year=anio,
            tipo_movimiento=TipoMovimiento.ENTRADA
        )
        
        dias_asistidos = asistencias.values('fecha').distinct().count()
        retardos = asistencias.filter(retardo=True).count()
        total_min_retardo = sum(asistencias.filter(retardo=True).values_list('minutos_retardo', flat=True))
        
        # Contar permisos aprobados
        permisos_dias = SolicitudPermiso.objects.filter(
            empleado=empleado,
            fecha_inicio__lte=fecha_fin_mes,
            fecha_fin__gte=fecha_inicio_mes,
            estado__in=[EstadoSolicitud.APROBADO_JEFE, EstadoSolicitud.APROBADO_GERENCIA]
        ).count()
        
        # Calcular días laborales esperados
        horario = obtener_horario_esperado(empleado, fecha_inicio_mes)
        if horario['tipo_sistema'] == 'TURNO_24H':
            # Turnos 24h: aproximadamente 15 turnos al mes
            dias_esperados = 15
        else:
            # Contar días laborales (lunes a viernes por defecto)
            dias_esperados = 0
            fecha_temp = fecha_inicio_mes
            while fecha_temp <= fecha_fin_mes:
                if fecha_temp.weekday() < 5:  # Lunes a viernes
                    dias_esperados += 1
                fecha_temp += timedelta(days=1)
        
        faltas = dias_esperados - dias_asistidos - permisos_dias
        if faltas < 0:
            faltas = 0
        
        ws_resumen.cell(row=row_resumen, column=1, value=empleado.user.get_full_name())
        ws_resumen.cell(row=row_resumen, column=2, value=empleado.codigo_empleado)
        ws_resumen.cell(row=row_resumen, column=3, value=empleado.departamento.nombre if empleado.departamento else 'N/A')
        ws_resumen.cell(row=row_resumen, column=4, value=dias_asistidos)
        ws_resumen.cell(row=row_resumen, column=5, value=retardos)
        ws_resumen.cell(row=row_resumen, column=6, value=total_min_retardo)
        ws_resumen.cell(row=row_resumen, column=7, value=faltas)
        ws_resumen.cell(row=row_resumen, column=8, value=permisos_dias)
        
        for col in range(1, 9):
            ws_resumen.cell(row=row_resumen, column=col).border = border
        
        row_resumen += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 12
    ws_resumen.column_dimensions['C'].width = 20
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws_resumen.column_dimensions[col].width = 15
    
    # ===== HOJA 2: DETALLE DE ASISTENCIAS =====
    ws_detalle['A1'] = "DETALLE DE TODAS LAS ASISTENCIAS"
    ws_detalle['A1'].font = title_font
    ws_detalle.merge_cells('A1:G1')
    
    headers_detalle = ['Fecha', 'Empleado', 'Código', 'Tipo Movimiento', 'Hora', 'Retardo', 'Min. Retardo']
    for col, header in enumerate(headers_detalle, start=1):
        cell = ws_detalle.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Obtener todas las asistencias del mes
    asistencias_mes = Asistencia.objects.filter(
        fecha__month=mes,
        fecha__year=anio
    ).select_related('empleado', 'empleado__user').order_by('fecha', 'empleado__codigo_empleado', 'hora')
    
    row_detalle = 4
    for asist in asistencias_mes:
        ws_detalle.cell(row=row_detalle, column=1, value=asist.fecha.strftime('%d/%m/%Y'))
        ws_detalle.cell(row=row_detalle, column=2, value=asist.empleado.user.get_full_name())
        ws_detalle.cell(row=row_detalle, column=3, value=asist.empleado.codigo_empleado)
        ws_detalle.cell(row=row_detalle, column=4, value=asist.get_tipo_movimiento_display())
        ws_detalle.cell(row=row_detalle, column=5, value=asist.hora.strftime('%H:%M:%S'))
        ws_detalle.cell(row=row_detalle, column=6, value='Sí' if asist.retardo else 'No')
        ws_detalle.cell(row=row_detalle, column=7, value=asist.minutos_retardo if asist.retardo else 0)
        
        for col in range(1, 8):
            ws_detalle.cell(row=row_detalle, column=col).border = border
        
        row_detalle += 1
    
    # Ajustar anchos
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_detalle.column_dimensions[col].width = 20
    ws_detalle.column_dimensions['F'].width = 12
    ws_detalle.column_dimensions['G'].width = 15
    
    # ===== HOJA 3: RETARDOS Y FALTAS =====
    ws_retardos['A1'] = "EMPLEADOS CON RETARDOS Y FALTAS"
    ws_retardos['A1'].font = title_font
    ws_retardos.merge_cells('A1:F1')
    
    headers_retardos = ['Empleado', 'Código', 'Departamento', 'Total Retardos', 'Total Min.', 'Faltas']
    for col, header in enumerate(headers_retardos, start=1):
        cell = ws_retardos.cell(row=3, column=col, value=header)
        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row_retardos = 4
    for empleado in empleados:
        asistencias = Asistencia.objects.filter(
            empleado=empleado,
            fecha__month=mes,
            fecha__year=anio,
            tipo_movimiento=TipoMovimiento.ENTRADA
        )
        
        retardos = asistencias.filter(retardo=True)
        total_retardos = retardos.count()
        total_min = sum(retardos.values_list('minutos_retardo', flat=True))
        
        dias_asistidos = asistencias.values('fecha').distinct().count()
        
        # Calcular faltas
        horario = obtener_horario_esperado(empleado, fecha_inicio_mes)
        if horario['tipo_sistema'] == 'TURNO_24H':
            dias_esperados = 15
        else:
            dias_esperados = sum(1 for d in range(dias_mes) if date(anio, mes, d+1).weekday() < 5)
        
        faltas = dias_esperados - dias_asistidos
        if faltas < 0:
            faltas = 0
        
        # Solo incluir empleados con retardos o faltas
        if total_retardos > 0 or faltas > 0:
            ws_retardos.cell(row=row_retardos, column=1, value=empleado.user.get_full_name())
            ws_retardos.cell(row=row_retardos, column=2, value=empleado.codigo_empleado)
            ws_retardos.cell(row=row_retardos, column=3, value=empleado.departamento.nombre if empleado.departamento else 'N/A')
            ws_retardos.cell(row=row_retardos, column=4, value=total_retardos)
            ws_retardos.cell(row=row_retardos, column=5, value=total_min)
            ws_retardos.cell(row=row_retardos, column=6, value=faltas)
            
            for col in range(1, 7):
                ws_retardos.cell(row=row_retardos, column=col).border = border
            
            row_retardos += 1
    
    # Ajustar anchos
    ws_retardos.column_dimensions['A'].width = 30
    ws_retardos.column_dimensions['B'].width = 12
    ws_retardos.column_dimensions['C'].width = 20
    for col in ['D', 'E', 'F']:
        ws_retardos.column_dimensions[col].width = 15
    
    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
